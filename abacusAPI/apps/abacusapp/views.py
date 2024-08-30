import base64
import logging
import os

import openpyxl
from django.conf import settings
from django.core.files.storage import default_storage
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import render
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.views import APIView

from abacusAPI.apps.abacusapp.services import (
    calculate_portfolio_daily_value,
    calculate_portfolio_daily_values,
    generate_portfolio_plots,
)

from .filters import PortfolioAssetFilter
from .models import Asset, Deposit, Portfolio, PortfolioAsset, Price, Transaction
from .serializers import (
    AssetSerializer,
    DepositSerializer,
    PortfolioAssetSerializer,
    PortfolioDailyValueSerializer,
    PortfolioSerializer,
    PriceSerializer,
    TransactionSerializer,
)

logger = logging.getLogger("abacusapp")


class PortfolioViewSet(viewsets.ModelViewSet):
    queryset = Portfolio.objects.all()
    serializer_class = PortfolioSerializer

    @action(detail=True, methods=["get"])
    def daily_value(self, request, pk=None):
        portfolio = self.get_object()
        date = request.query_params.get("date")

        if not date:
            return Response(
                {"error": "Please provide a date."}, status=status.HTTP_400_BAD_REQUEST
            )

        daily_value = calculate_portfolio_daily_value(portfolio, date)
        serializer = PortfolioDailyValueSerializer(daily_value)

        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["get"])
    def daily_values(self, request, pk=None):
        portfolio = self.get_object()
        initial_date = request.query_params.get("fecha_inicio")
        end_date = request.query_params.get("fecha_fin")

        if not initial_date or not end_date:
            return Response(
                {"error": "Please provide both fecha_inicio and fecha_fin."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        daily_values = calculate_portfolio_daily_values(
            portfolio, initial_date, end_date
        )
        serializer = PortfolioDailyValueSerializer(daily_values, many=True)

        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["get"])
    def plot(self, request, pk=None):
        portfolio = self.get_object()
        initial_date = request.query_params.get("fecha_inicio")
        end_date = request.query_params.get("fecha_fin")

        if not initial_date or not end_date:
            return Response(
                {"error": "Please provide both fecha_inicio and fecha_fin."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        image_base64 = generate_portfolio_plots(portfolio, initial_date, end_date)
        if not image_base64:
            return Response(
                {"error": "No data available for the given date range."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Return the image as a response
        return HttpResponse(base64.b64decode(image_base64), content_type="image/png")


class AssetViewSet(viewsets.ModelViewSet):
    queryset = Asset.objects.all()
    serializer_class = AssetSerializer


class PriceViewSet(viewsets.ModelViewSet):
    queryset = Price.objects.all().order_by("date")
    serializer_class = PriceSerializer
    pagination_class = PageNumberPagination


class PortfolioAssetViewSet(viewsets.ModelViewSet):
    queryset = PortfolioAsset.objects.all()
    serializer_class = PortfolioAssetSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = PortfolioAssetFilter


class DepositViewSet(viewsets.ModelViewSet):
    queryset = Deposit.objects.all()
    serializer_class = DepositSerializer


class TransactionViewSet(viewsets.ModelViewSet):
    queryset = Transaction.objects.all()
    serializer_class = TransactionSerializer

    def create(self, request, *args, **kwargs):
        """
        Catch errors when creating a related PortfolioAsset
        """
        try:
            return super().create(request, *args, **kwargs)
        except ValidationError as e:
            return Response({"error": e.detail}, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        """
        Catch errors when updating a related PortfolioAsset
        """
        try:
            return super().update(request, *args, **kwargs)
        except ValidationError as e:
            return Response({"error": e.detail}, status=status.HTTP_400_BAD_REQUEST)


class UploadExcelView(APIView):
    """
    View to be able to load the .xlsx setup file
    """

    def get(self, request, *args, **kwargs):
        return render(request, "upload_excel.html")

    def post(self, request, *args, **kwargs):
        logger.info("Started File Upload")

        # Handle the file upload
        if "file" not in request.FILES:
            return Response(
                {"error": "No file uploaded"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        f = request.FILES["file"]
        if not f.name.endswith(".xlsx"):
            return Response(
                {"error": "File is not in .xlsx format"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        logger.info(f"Received File Correctly [file={f.name}]")

        # save the file
        file_path = default_storage.save("temp/" + f.name, f)
        file_full_path = os.path.join(settings.MEDIA_ROOT, file_path)

        # TODO: The file logic could be split to helper functions
        # to have a cleaner code structure
        try:
            # Read the file
            workbook = openpyxl.load_workbook(file_full_path)

            # Process "weights" sheet
            weights_sheet = workbook["weights"]

            # Portfolio names are in the header row starting from the third column
            portfolio_names = [cell.value for cell in weights_sheet[1][2:]]
            portfolios = {}
            # Create portfolios if they don't exist
            for name in portfolio_names:
                logger.debug("Creating Portfolio [name=%s]", name)
                portfolio, _ = Portfolio.objects.get_or_create(name=name)
                portfolios[name] = portfolio

            # this is to avoid getting/creating Assets multiple times
            assets_by_name = {}

            # Process each row for the assets and weights
            for row in weights_sheet.iter_rows(min_row=2, values_only=True):
                date, asset_name, *weights = row

                # stop when rows are empty
                if not date or not asset_name:
                    logger.info("Done processing Portfolios")
                    break

                # create the asset if it doesn't exist
                logger.debug("Retrieving Asset [name=%s]", asset_name)
                assets_by_name[asset_name] = Asset.objects.get_or_create(
                    name=asset_name,
                )[0]

                # then the weights and amounts
                # TODO: this could also be a transaction so it's faster
                for portfolio_name, weight in zip(portfolio_names, weights):
                    logger.debug(
                        "Creating PortfolioAsset [p=%s, a=%s, d=%s, w=%s]",
                        portfolio_name,
                        asset_name,
                        date,
                        weight,
                    )
                    PortfolioAsset.objects.update_or_create(
                        portfolio=portfolios[portfolio_name],
                        asset=assets_by_name[asset_name],
                        weight=weight,
                    )

            # Process "Precios" sheet
            prices_sheet = workbook["Precios"]

            # Get the asset names from the header row
            asset_names = [cell.value for cell in prices_sheet[1][1:]]

            # First get/create the assets and save them
            for name in asset_names:
                if name in assets_by_name:
                    # skip if we saw it already on the previous step
                    continue
                logger.debug("Retrieving Asset [name=%s]", name)
                assets_by_name[name] = Asset.objects.get_or_create(
                    name=name,
                )[0]

            # then create or update their prices in bulk
            new_prices = []
            prices_to_update = []

            for row in prices_sheet.iter_rows(min_row=2, values_only=True):
                date = row[0]
                for asset_name, price in zip(asset_names, row[1:]):
                    logger.debug(
                        "Preparing Asset Price [asset=%s, price=%s, date=%s]",
                        asset_name,
                        price,
                        date,
                    )

                    asset = assets_by_name[asset_name]
                    price_obj = Price.objects.filter(
                        asset=asset,
                        date=date,
                    ).first()

                    if price_obj:
                        price_obj.value = price
                        prices_to_update.append(price_obj)
                    else:
                        new_prices.append(
                            Price(
                                asset=asset,
                                date=date,
                                price=price,
                            ),
                        )

            # transaction to ensure atomicity
            with transaction.atomic():
                # Bulk create new prices
                if new_prices:
                    logger.debug("Creating Asset Prices (%d)", len(new_prices))
                    Price.objects.bulk_create(new_prices)

                # Bulk update existing prices
                if prices_to_update:
                    logger.debug("Updating Asset Prices (%d)", len(prices_to_update))
                    Price.objects.bulk_update(prices_to_update, ["price"])

        except Exception as e:
            logger.exception("File Upload Failed")
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        finally:
            # Always clean up the file
            os.remove(file_full_path)

        return Response(
            {"status": "File processed successfully"},
            status=status.HTTP_200_OK,
        )
